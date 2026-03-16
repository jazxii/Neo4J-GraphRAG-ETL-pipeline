"""
WCAG 2.2 Comprehensive Scraper → CSV / Excel

Scrapes every Success Criterion's official W3C "Understanding" page and
extracts ALL structured data into a flat, scalable CSV (and optional Excel).

Data Points per Success Criterion:
  ┌─────────────────────────────────────────────────────────────────┐
  │  CORE METADATA                                                  │
  │  ref_id · title · level · principle · guideline · url           │
  ├─────────────────────────────────────────────────────────────────┤
  │  IN BRIEF                                                       │
  │  in_brief_goal · in_brief_what_to_do · in_brief_why_important  │
  ├─────────────────────────────────────────────────────────────────┤
  │  UNDERSTANDING                                                  │
  │  intent (full text) · benefits · examples                       │
  ├─────────────────────────────────────────────────────────────────┤
  │  TECHNIQUES                                                     │
  │  sufficient_techniques · advisory_techniques · failure_techniques│
  ├─────────────────────────────────────────────────────────────────┤
  │  TESTING & REFERENCES                                           │
  │  test_rules · related_resources · key_terms                     │
  ├─────────────────────────────────────────────────────────────────┤
  │  RELATIONSHIPS                                                  │
  │  related_scs (cross-referenced criteria)                        │
  ├─────────────────────────────────────────────────────────────────┤
  │  DERIVED METADATA                                               │
  │  wcag_version · disability_impact · input_types_affected        │
  │  automatable                                                    │
  ├─────────────────────────────────────────────────────────────────┤
  │  PROVENANCE                                                     │
  │  source_url · scrape_timestamp                                  │
  └─────────────────────────────────────────────────────────────────┘

Scalability:
  - Add new data sources by extending DATA_SOURCES list
  - Each source defines: name, base_url, extractor function
  - CSV columns are auto-extended when new sources add new fields
  - Re-run is idempotent — merges by ref_id

Usage:
  pip install requests beautifulsoup4 openpyxl
  python 00_scrape_wcag_to_csv.py

Output:
  wcag_22_all_criteria.csv
  wcag_22_all_criteria.xlsx  (if openpyxl installed)
  wcag_22_guidelines_enriched.json  (updated enriched JSON)
"""

import csv
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone

try:
    import requests
    from bs4 import BeautifulSoup, Tag
except ImportError:
    print("Install dependencies:  pip install requests beautifulsoup4")
    sys.exit(1)

# ── Logging ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-7s │ %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("wcag_scraper")

# ── Config ──
INPUT_JSON = "wcag_22_guidelines.json"
OUTPUT_CSV = "wcag_22_all_criteria.csv"
OUTPUT_XLSX = "wcag_22_all_criteria.xlsx"
OUTPUT_JSON = "wcag_22_guidelines_enriched.json"
RATE_LIMIT = 1.2  # seconds between requests (be polite to W3C)

# ════════════════════════════════════════════════════════════
# DERIVED METADATA MAPS (no scraping needed)
# ════════════════════════════════════════════════════════════

WCAG_21_NEW = {
    "1.3.4", "1.3.5", "1.3.6", "1.4.10", "1.4.11", "1.4.12", "1.4.13",
    "2.1.4", "2.5.1", "2.5.2", "2.5.3", "2.5.4", "2.5.5", "2.5.6",
    "4.1.3",
}
WCAG_22_NEW = {
    "2.4.11", "2.4.12", "2.4.13",
    "2.5.7", "2.5.8",
    "3.2.6", "3.3.7", "3.3.8", "3.3.9",
}

DISABILITY_IMPACT_MAP = {
    "1.1": ["blindness", "low_vision", "deafblindness", "cognitive", "learning"],
    "1.2": ["deafness", "hard_of_hearing", "blindness", "deafblindness", "cognitive", "learning"],
    "1.3": ["blindness", "low_vision", "cognitive", "learning", "mobility"],
    "1.4": ["low_vision", "color_blindness", "cognitive", "hard_of_hearing"],
    "2.1": ["blindness", "mobility", "motor_impairment"],
    "2.2": ["blindness", "low_vision", "cognitive", "learning", "motor_impairment", "deafness"],
    "2.3": ["photosensitive_epilepsy", "vestibular_disorders"],
    "2.4": ["blindness", "low_vision", "cognitive", "learning", "motor_impairment"],
    "2.5": ["mobility", "motor_impairment", "tremor", "cognitive"],
    "3.1": ["cognitive", "learning", "deafness", "blindness"],
    "3.2": ["blindness", "cognitive", "learning", "low_vision"],
    "3.3": ["blindness", "cognitive", "learning", "motor_impairment", "low_vision"],
    "4.1": ["blindness", "low_vision", "cognitive", "motor_impairment"],
}

FULLY_AUTOMATABLE = {
    "1.3.1", "1.3.2", "1.4.3", "1.4.6", "1.4.10", "1.4.12",
    "2.4.1", "2.4.2", "3.1.1", "3.1.2", "4.1.1", "4.1.2",
}
PARTIALLY_AUTOMATABLE = {
    "1.1.1", "1.2.1", "1.2.2", "1.2.3", "1.2.4", "1.2.5",
    "1.3.3", "1.3.4", "1.3.5", "1.4.1", "1.4.4", "1.4.5",
    "1.4.11", "1.4.13",
    "2.1.1", "2.1.2", "2.1.4", "2.2.1", "2.2.2",
    "2.4.3", "2.4.4", "2.4.6", "2.4.7", "2.4.11",
    "2.5.1", "2.5.2", "2.5.3", "2.5.7", "2.5.8",
    "3.2.1", "3.2.2", "3.2.6", "3.3.1", "3.3.2", "3.3.7", "3.3.8",
    "4.1.3",
}


def get_wcag_version(ref_id: str) -> str:
    if ref_id in WCAG_22_NEW: return "2.2"
    if ref_id in WCAG_21_NEW: return "2.1"
    return "2.0"


def get_automatable(ref_id: str) -> str:
    if ref_id in FULLY_AUTOMATABLE: return "full"
    if ref_id in PARTIALLY_AUTOMATABLE: return "partial"
    return "manual"


def get_disability_impact(ref_id: str) -> list[str]:
    guideline = ".".join(ref_id.split(".")[:2])
    return DISABILITY_IMPACT_MAP.get(guideline, ["general"])


def classify_input_types(ref_id: str, description: str) -> list[str]:
    types = set()
    desc_lower = description.lower()
    guideline = ".".join(ref_id.split(".")[:2])
    if guideline == "2.1": types.add("keyboard")
    elif guideline == "2.5": types.update(["pointer", "touch", "gesture"])
    kw_map = {
        "keyboard": ["keyboard", "focus", "tab", "key shortcut"],
        "pointer": ["pointer", "click", "mouse", "drag", "target size"],
        "touch": ["touch", "gesture", "swipe", "pinch"],
        "voice": ["voice", "speech"],
        "visual": ["color", "contrast", "visual", "text", "image", "visible", "reflow"],
        "auditory": ["audio", "captions", "sound", "sign language"],
    }
    for input_type, keywords in kw_map.items():
        if any(kw in desc_lower for kw in keywords):
            types.add(input_type)
    return sorted(types) if types else ["visual"]


# ════════════════════════════════════════════════════════════
# HTML SECTION EXTRACTORS
# ════════════════════════════════════════════════════════════

def fetch_soup(url: str) -> BeautifulSoup | None:
    """Fetch a URL and return parsed soup."""
    try:
        resp = requests.get(url, timeout=20, headers={
            "User-Agent": "WCAG-KnowledgeGraph-Scraper/1.0 (accessibility research)"
        })
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        log.warning("  Failed to fetch %s: %s", url, e)
        return None


def _get_section_tag(soup: BeautifulSoup, section_id: str) -> Tag | None:
    """Find a <section> or heading by its id attribute."""
    # Try <section id="...">
    tag = soup.find("section", id=section_id)
    if tag:
        return tag
    # Try <h2 id="..."> or <h3 id="...">
    tag = soup.find(id=section_id)
    return tag


def _extract_section_text(soup: BeautifulSoup, section_id: str) -> str:
    """Extract all text from a section by its id, stopping at next same-level heading."""
    section = _get_section_tag(soup, section_id)
    if not section:
        return ""

    # If it's a <section> element, get all text inside
    if section.name == "section":
        return section.get_text(separator="\n", strip=True)

    # Otherwise it's a heading — collect siblings until next heading of same or higher level
    heading_level = int(section.name[1]) if section.name and section.name[0] == "h" else 2
    parts = []
    for sib in section.find_next_siblings():
        if sib.name and re.match(r"^h[2-9]$", sib.name):
            sib_level = int(sib.name[1])
            if sib_level <= heading_level:
                break
        text = sib.get_text(separator="\n", strip=True)
        if text:
            parts.append(text)
    return "\n".join(parts)


def extract_in_brief(soup: BeautifulSoup) -> dict:
    """Extract the In Brief section → {goal, what_to_do, why_important}.

    W3C structure:
        <section id="brief">
          <h2>In Brief</h2>
          <dl>
            <dt>Goal</dt><dd>…</dd>
            <dt>What to do</dt><dd>…</dd>
            <dt>Why it's important</dt><dd>…</dd>
          </dl>
        </section>
    """
    brief = {"goal": "", "what_to_do": "", "why_important": ""}

    section = _get_section_tag(soup, "brief")
    if not section:
        return brief

    container = section if section.name == "section" else section.parent

    # Primary approach: <dl> with <dt>/<dd> pairs
    dl = (container or section).find("dl")
    if dl:
        for dt in dl.find_all("dt"):
            label = dt.get_text(strip=True).lower()
            dd = dt.find_next_sibling("dd")
            value = dd.get_text(separator=" ", strip=True) if dd else ""
            if "goal" in label:
                brief["goal"] = value
            elif "what to do" in label:
                brief["what_to_do"] = value
            elif "why" in label:
                brief["why_important"] = value
        return brief

    # Fallback: <strong> label approach (older pages)
    for strong in (container or soup).find_all("strong"):
        label = strong.get_text(strip=True).lower()
        value_parts = []
        for sib in strong.next_siblings:
            if isinstance(sib, Tag) and sib.name == "strong":
                break
            text = sib.get_text(strip=True) if isinstance(sib, Tag) else str(sib).strip()
            if text:
                value_parts.append(text)
        value = " ".join(value_parts).strip()
        if "goal" in label:
            brief["goal"] = value
        elif "what to do" in label:
            brief["what_to_do"] = value
        elif "why" in label:
            brief["why_important"] = value

    return brief


def extract_intent(soup: BeautifulSoup) -> str:
    """Extract the Intent section as clean text."""
    text = _extract_section_text(soup, "intent")
    # Clean up: remove "Intent" heading text if it starts with it
    text = re.sub(r"^Intent\s*\n?", "", text).strip()
    return text


def extract_benefits(soup: BeautifulSoup) -> list[str]:
    """Extract the Benefits section as a list of benefit strings."""
    section = _get_section_tag(soup, "benefits")
    if not section:
        return []

    benefits = []
    # Find the <ul> within the section
    container = section if section.name == "section" else section.parent
    if not container:
        container = section

    ul = container.find("ul") if container.name == "section" else None
    if not ul:
        # Try next sibling
        for sib in (section.find_next_siblings() if section.name != "section" else []):
            if isinstance(sib, Tag) and sib.name == "ul":
                ul = sib
                break
            if isinstance(sib, Tag) and sib.name and re.match(r"^h[2-3]$", sib.name):
                break

    if ul:
        for li in ul.find_all("li", recursive=False):
            text = li.get_text(separator=" ", strip=True)
            if text:
                benefits.append(text)
    else:
        # Fallback: get paragraph text
        text = _extract_section_text(soup, "benefits")
        text = re.sub(r"^Benefits\s*\n?", "", text).strip()
        if text:
            benefits = [line.strip() for line in text.split("\n") if line.strip()]

    return benefits


def extract_examples(soup: BeautifulSoup) -> list[dict]:
    """Extract Examples section as list of {title, description}.

    W3C structure:
        <section id="examples">
          <h2>Examples</h2>
          <dl>
            <dt>Example title</dt>
            <dd>Example description…</dd>
            …
          </dl>
        </section>
    """
    section = _get_section_tag(soup, "examples")
    if not section:
        return []

    examples = []
    container = section if section.name == "section" else None

    # Primary approach: <dl> with <dt>/<dd> pairs
    search_root = container if container else section
    dl = search_root.find("dl") if container else None
    if not dl:
        for sib in (section.find_next_siblings() if not container else []):
            if isinstance(sib, Tag) and sib.name == "dl":
                dl = sib
                break
            if isinstance(sib, Tag) and sib.name and re.match(r"^h[2-3]$", sib.name):
                break

    if dl:
        for dt in dl.find_all("dt"):
            title = dt.get_text(strip=True)
            dd = dt.find_next_sibling("dd")
            description = dd.get_text(separator=" ", strip=True) if dd else ""
            if title or description:
                examples.append({"title": title, "description": description})
        return examples

    # Fallback: <ul>/<ol> lists (some older pages may use this)
    lists = []
    if container:
        lists = container.find_all(["ul", "ol"], recursive=True)
    else:
        for sib in section.find_next_siblings():
            if isinstance(sib, Tag) and sib.name in ("ul", "ol"):
                lists.append(sib)
                break
            if isinstance(sib, Tag) and sib.name and re.match(r"^h[2-3]$", sib.name):
                break

    for lst in lists:
        for li in lst.find_all("li", recursive=False):
            strong = li.find("strong")
            title = strong.get_text(strip=True) if strong else ""
            desc = li.get_text(separator=" ", strip=True)
            if strong and title:
                desc = desc.replace(title, "", 1).strip()
            if desc:
                examples.append({"title": title, "description": desc})

    # Last fallback: paragraphs
    if not examples:
        text = _extract_section_text(soup, "examples")
        text = re.sub(r"^Examples\s*\n?", "", text).strip()
        if text:
            for para in text.split("\n\n"):
                para = para.strip()
                if para:
                    examples.append({"title": "", "description": para})

    return examples


def extract_techniques(soup: BeautifulSoup) -> dict:
    """Extract Sufficient, Advisory, and Failure techniques."""
    result = {"sufficient": [], "advisory": [], "failures": []}

    techniques_section = _get_section_tag(soup, "techniques")
    if not techniques_section:
        return result

    # Determine the container to search within
    if techniques_section.name == "section":
        container = techniques_section
    else:
        # Search everything after the Techniques heading
        container = soup

    # Find each sub-section
    for sub_id, key in [
        ("sufficient", "sufficient"),
        ("advisory", "advisory"),
        ("failures", "failures"),
        ("failure", "failures"),
    ]:
        sub = container.find("section", id=sub_id)
        if not sub:
            sub = container.find(id=sub_id)

        search_area = sub if sub else container

        # Find technique links within this subsection
        for a_tag in search_area.find_all("a", href=True) if search_area else []:
            href = a_tag.get("href", "")
            text = a_tag.get_text(separator=" ", strip=True)

            # Match technique URLs
            if "/Techniques/" not in href and "/techniques/" not in href:
                continue
            if not text:
                continue

            # Extract technique ID
            match = re.search(r"/([A-Z]+\d+)/?(?:\.\w+)?$", href)
            tech_id = match.group(1) if match else ""
            if not tech_id:
                continue

            tech_prefix = re.match(r"^([A-Z]+)", tech_id)
            tech_type_map = {
                "G": "general", "H": "html", "C": "css", "ARIA": "aria",
                "PDF": "pdf", "F": "failure", "SCR": "script",
                "SM": "smil", "SVR": "server",
            }
            technology = tech_type_map.get(tech_prefix.group(1), "general") if tech_prefix else "general"

            full_url = href if href.startswith("http") else f"https://www.w3.org{href}"

            entry = {
                "id": tech_id,
                "title": text,
                "url": full_url,
                "technology": technology,
            }

            # Classify: if we found this in a specific sub-section use that
            if sub:
                if key not in result:
                    result[key] = []
                result[key].append(entry)
            else:
                # Fallback: failures start with F
                if tech_id.startswith("F"):
                    result["failures"].append(entry)
                else:
                    # Try to detect from parent context
                    parent_text = ""
                    parent = a_tag.find_parent(["section", "div"])
                    if parent:
                        h = parent.find(re.compile(r"^h[3-5]$"))
                        parent_text = h.get_text().lower() if h else ""
                    if "advisory" in parent_text:
                        result["advisory"].append(entry)
                    elif "failur" in parent_text:
                        result["failures"].append(entry)
                    else:
                        result["sufficient"].append(entry)

    # Deduplicate each list by tech_id
    for key in result:
        seen = set()
        deduped = []
        for t in result[key]:
            if t["id"] not in seen:
                seen.add(t["id"])
                deduped.append(t)
        result[key] = deduped

    return result


def extract_test_rules(soup: BeautifulSoup) -> list[dict]:
    """Extract ACT Test Rules."""
    rules = []
    section = _get_section_tag(soup, "test-rules")
    if not section:
        return rules

    container = section if section.name == "section" else section.parent
    search_area = container if container else section

    for a_tag in search_area.find_all("a", href=True) if search_area else []:
        href = a_tag.get("href", "")
        text = a_tag.get_text(strip=True)
        if href and ("act/rules" in href or "standards-guidelines/act" in href):
            full_url = href if href.startswith("http") else f"https://www.w3.org{href}"
            rules.append({"title": text, "url": full_url})

    # Also try siblings if section is a heading
    if not rules and section.name and section.name.startswith("h"):
        for sib in section.find_next_siblings():
            if isinstance(sib, Tag) and sib.name and re.match(r"^h[2-3]$", sib.name):
                break
            for a_tag in (sib.find_all("a", href=True) if isinstance(sib, Tag) else []):
                href = a_tag.get("href", "")
                text = a_tag.get_text(strip=True)
                if href and ("act/rules" in href or "standards-guidelines/act" in href):
                    full_url = href if href.startswith("http") else f"https://www.w3.org{href}"
                    rules.append({"title": text, "url": full_url})

    return rules


def extract_related_resources(soup: BeautifulSoup) -> list[dict]:
    """Extract Related Resources section."""
    resources = []
    section = _get_section_tag(soup, "resources")
    if not section:
        return resources

    search_area = section if section.name == "section" else None
    if not search_area:
        search_area = section.parent if section.parent else section

    for a_tag in search_area.find_all("a", href=True):
        href = a_tag.get("href", "")
        text = a_tag.get_text(strip=True)
        if href and text and not href.startswith("#") and not href.startswith("mailto:"):
            # Skip navigation links
            if "WAI/WCAG22/Understanding" in href and text in ("All Understanding Docs",):
                continue
            full_url = href if href.startswith("http") else f"https://www.w3.org{href}"
            resources.append({"title": text, "url": full_url})

    return resources


def extract_related_scs(soup: BeautifulSoup, own_ref_id: str) -> list[str]:
    """Find cross-referenced success criteria from Intent and Benefits sections."""
    related = set()

    # Search in Intent + Benefits sections
    for section_id in ("intent", "benefits"):
        text = _extract_section_text(soup, section_id)
        # Find SC patterns like "1.4.3" or "Success Criterion 2.4.7"
        for match in re.finditer(r"\b(\d+\.\d+\.\d+)\b", text):
            found_id = match.group(1)
            if found_id != own_ref_id:
                related.add(found_id)

    # Also check technique links for cross-referenced SCs
    for a_tag in soup.find_all("a", href=True):
        href = a_tag.get("href", "")
        text = a_tag.get_text(strip=True)
        for match in re.finditer(r"\b(\d+\.\d+\.\d+)\b", text):
            found_id = match.group(1)
            if found_id != own_ref_id:
                related.add(found_id)

    return sorted(related)


def extract_key_terms(soup: BeautifulSoup) -> list[dict]:
    """Extract key terms definitions."""
    terms = []
    section = _get_section_tag(soup, "key-terms")
    if not section:
        return terms

    container = section if section.name == "section" else section.parent
    if not container:
        return terms

    # Key terms are typically in <dt>/<dd> pairs or <h3>+<p> pairs
    for dt in container.find_all("dt"):
        term = dt.get_text(strip=True)
        dd = dt.find_next_sibling("dd")
        definition = dd.get_text(separator=" ", strip=True) if dd else ""
        if term:
            terms.append({"term": term, "definition": definition})

    return terms


# ════════════════════════════════════════════════════════════
# UNDERSTANDING PAGE URL BUILDER
# ════════════════════════════════════════════════════════════

def get_understanding_url(sc_url: str) -> str:
    """Derive the Understanding page URL from the criterion's spec URL."""
    slug = sc_url.split("#")[-1] if "#" in sc_url else ""
    if slug:
        return f"https://www.w3.org/WAI/WCAG22/Understanding/{slug}.html"
    return ""


# ════════════════════════════════════════════════════════════
# MAIN SCRAPER
# ════════════════════════════════════════════════════════════

def scrape_criterion(sc: dict, principle_title: str, guideline_title: str) -> dict:
    """Scrape all data for a single success criterion."""
    ref_id = sc["ref_id"]
    understanding_url = get_understanding_url(sc["url"])

    log.info("  Scraping %s: %s", ref_id, sc["title"])

    row = {
        # ── Core metadata ──
        "ref_id": ref_id,
        "title": sc["title"],
        "level": sc["level"],
        "description": sc.get("description", ""),
        "url": sc.get("url", ""),
        "principle": principle_title,
        "guideline": guideline_title,

        # ── In Brief ──
        "in_brief_goal": "",
        "in_brief_what_to_do": "",
        "in_brief_why_important": "",

        # ── Understanding ──
        "intent": "",
        "benefits": "[]",
        "examples": "[]",

        # ── Techniques ──
        "sufficient_techniques": "[]",
        "advisory_techniques": "[]",
        "failure_techniques": "[]",

        # ── Testing & References ──
        "test_rules": "[]",
        "related_resources": "[]",
        "key_terms": "[]",

        # ── Relationships ──
        "related_scs": "[]",

        # ── Derived metadata ──
        "wcag_version": get_wcag_version(ref_id),
        "disability_impact": json.dumps(get_disability_impact(ref_id)),
        "input_types_affected": json.dumps(classify_input_types(ref_id, sc.get("description", ""))),
        "automatable": get_automatable(ref_id),

        # ── Provenance ──
        "source_url": understanding_url,
        "scrape_timestamp": datetime.now(timezone.utc).isoformat(),
    }

    if not understanding_url:
        log.warning("    ⚠ No Understanding URL for %s", ref_id)
        return row

    soup = fetch_soup(understanding_url)
    if not soup:
        log.warning("    ⚠ Could not fetch %s", understanding_url)
        return row

    # ── Extract each section ──
    in_brief = extract_in_brief(soup)
    row["in_brief_goal"] = in_brief.get("goal", "")
    row["in_brief_what_to_do"] = in_brief.get("what_to_do", "")
    row["in_brief_why_important"] = in_brief.get("why_important", "")

    row["intent"] = extract_intent(soup)

    benefits = extract_benefits(soup)
    row["benefits"] = json.dumps(benefits, ensure_ascii=False)

    examples = extract_examples(soup)
    row["examples"] = json.dumps(examples, ensure_ascii=False)

    techniques = extract_techniques(soup)
    row["sufficient_techniques"] = json.dumps(techniques.get("sufficient", []), ensure_ascii=False)
    row["advisory_techniques"] = json.dumps(techniques.get("advisory", []), ensure_ascii=False)
    row["failure_techniques"] = json.dumps(techniques.get("failures", []), ensure_ascii=False)

    test_rules = extract_test_rules(soup)
    row["test_rules"] = json.dumps(test_rules, ensure_ascii=False)

    related_resources = extract_related_resources(soup)
    row["related_resources"] = json.dumps(related_resources, ensure_ascii=False)

    key_terms = extract_key_terms(soup)
    row["key_terms"] = json.dumps(key_terms, ensure_ascii=False)

    related_scs = extract_related_scs(soup, ref_id)
    row["related_scs"] = json.dumps(related_scs)

    # Log coverage
    suf_count = len(techniques.get("sufficient", []))
    adv_count = len(techniques.get("advisory", []))
    fail_count = len(techniques.get("failures", []))
    log.info(
        "    ✅ intent=%d chars, brief=%s, benefits=%d, examples=%d, "
        "techniques=%d/%d/%d, test_rules=%d, related=%d",
        len(row["intent"]),
        "yes" if any(in_brief.values()) else "no",
        len(benefits), len(examples),
        suf_count, adv_count, fail_count,
        len(test_rules), len(related_scs),
    )

    time.sleep(RATE_LIMIT)
    return row


def write_csv(rows: list[dict], filepath: str):
    """Write rows to CSV."""
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    log.info("Written %d rows to %s", len(rows), filepath)


def write_xlsx(rows: list[dict], filepath: str):
    """Write rows to Excel (if openpyxl is available)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        log.info("openpyxl not installed — skipping Excel output")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "WCAG 2.2 Criteria"

    if not rows:
        wb.save(filepath)
        return

    # Header
    fieldnames = list(rows[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width (approximate)
    for col_idx, field in enumerate(fieldnames, 1):
        max_len = len(field)
        for row in rows[:5]:
            val_len = len(str(row.get(field, "")))
            max_len = max(max_len, min(val_len, 60))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(filepath)
    log.info("Written %d rows to %s", len(rows), filepath)


def update_enriched_json(rows: list[dict], input_json: str, output_json: str):
    """Merge scraped data back into the enriched JSON structure."""
    with open(input_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Build lookup from scraped rows
    lookup = {r["ref_id"]: r for r in rows}

    for principle in data:
        for guideline in principle.get("guidelines", []):
            for sc in guideline.get("success_criteria", []):
                ref_id = sc["ref_id"]
                row = lookup.get(ref_id)
                if not row:
                    continue

                # Merge scraped data into SC
                sc["in_brief"] = {
                    "goal": row.get("in_brief_goal", ""),
                    "what_to_do": row.get("in_brief_what_to_do", ""),
                    "why_important": row.get("in_brief_why_important", ""),
                }
                sc["intent"] = row.get("intent", "")
                sc["benefits"] = json.loads(row.get("benefits", "[]"))
                sc["examples"] = json.loads(row.get("examples", "[]"))
                sc["techniques"] = {
                    "sufficient": json.loads(row.get("sufficient_techniques", "[]")),
                    "advisory": json.loads(row.get("advisory_techniques", "[]")),
                    "failures": json.loads(row.get("failure_techniques", "[]")),
                }
                sc["test_rules"] = json.loads(row.get("test_rules", "[]"))
                sc["related_scs"] = json.loads(row.get("related_scs", "[]"))
                sc["related_resources"] = json.loads(row.get("related_resources", "[]"))
                sc["key_terms"] = json.loads(row.get("key_terms", "[]"))
                sc["wcag_version"] = row.get("wcag_version", "2.0")
                sc["automatable"] = row.get("automatable", "manual")
                sc["disability_impact"] = json.loads(row.get("disability_impact", "[]"))
                sc["input_types_affected"] = json.loads(row.get("input_types_affected", "[]"))

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    log.info("Updated enriched JSON → %s", output_json)


# ════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════

def main():
    log.info("═" * 62)
    log.info("  WCAG 2.2 Comprehensive Scraper → CSV / Excel")
    log.info("═" * 62)

    # Load base JSON for SC list
    if not os.path.isfile(INPUT_JSON):
        log.error("Source file not found: %s", INPUT_JSON)
        sys.exit(1)

    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)

    log.info("Loaded %d principles from %s", len(data), INPUT_JSON)

    all_rows = []
    total = 0
    t_start = time.time()

    for principle in data:
        log.info("\nPrinciple %s: %s", principle["ref_id"], principle["title"])

        for guideline in principle.get("guidelines", []):
            log.info("  Guideline %s: %s", guideline["ref_id"], guideline["title"])

            for sc in guideline.get("success_criteria", []):
                row = scrape_criterion(sc, principle["title"], guideline["title"])
                all_rows.append(row)
                total += 1

    elapsed = round(time.time() - t_start, 1)
    log.info("\n" + "═" * 62)
    log.info("  Scraped %d criteria in %.1fs", total, elapsed)
    log.info("═" * 62)

    # ── Coverage report ──
    log.info("\n  COVERAGE REPORT:")
    fields_to_check = {
        "in_brief_goal": lambda v: bool(v),
        "intent": lambda v: len(v) > 50,
        "benefits": lambda v: len(json.loads(v)) > 0,
        "examples": lambda v: len(json.loads(v)) > 0,
        "sufficient_techniques": lambda v: len(json.loads(v)) > 0,
        "advisory_techniques": lambda v: len(json.loads(v)) > 0,
        "failure_techniques": lambda v: len(json.loads(v)) > 0,
        "test_rules": lambda v: len(json.loads(v)) > 0,
        "related_scs": lambda v: len(json.loads(v)) > 0,
    }
    for field, check in fields_to_check.items():
        has_data = sum(1 for r in all_rows if check(r.get(field, "")))
        pct = has_data / total * 100 if total else 0
        bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
        log.info("    %-25s %s %d/%d (%.0f%%)", field, bar, has_data, total, pct)

    # ── Write outputs ──
    write_csv(all_rows, OUTPUT_CSV)
    write_xlsx(all_rows, OUTPUT_XLSX)
    update_enriched_json(all_rows, INPUT_JSON, OUTPUT_JSON)

    log.info("\n✅ Done! Files written:")
    log.info("   %s", OUTPUT_CSV)
    if os.path.isfile(OUTPUT_XLSX):
        log.info("   %s", OUTPUT_XLSX)
    log.info("   %s  (enriched JSON updated)", OUTPUT_JSON)


if __name__ == "__main__":
    main()
